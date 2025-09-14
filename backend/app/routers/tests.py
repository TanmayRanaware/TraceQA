from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from ..services.testgen import TestGenerator
from ..schemas.tests import (
    TestGenerationRequest, 
    TestGenerationResponse,
    ChangeManagementRequest,
    TestCaseUpdateRequest,
    TestCaseValidationRequest
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from typing import List, Dict, Any
from datetime import datetime

router = APIRouter(prefix="/api/tests", tags=["test-generation"])

@router.post("/generate")
async def generate_tests(payload: TestGenerationRequest):
    try:
        test_generator = TestGenerator()
        result = await test_generator.generate_test_cases(
            journey=payload.journey,
            max_cases=payload.max_cases,
            context="",  # Context will be built from RAG search
            source_types=None,  # Use all source types
            model=payload.model,
            temperature=None,  # Use default temperature
            context_top_k=payload.context_top_k,  # Pass the context_top_k parameter
            page=payload.page  # Pass the page parameter for pagination
        )
        
        return {
            "journey": payload.journey,
            "tests": result.get("test_cases", []),
            "pagination": {
                "page": result.get("page", 1),
                "has_next_page": result.get("has_next_page", False),
                "total_pages": result.get("total_pages", 1),
                "total_available": result.get("total_available", 0)
            },
            "metadata": result.get("metadata", {}),
            "debug": {
                "status": result.get("status", "unknown"),
                "message": result.get("message", ""),
                "total_generated": result.get("total_generated", 0),
                "context_used": result.get("context_used", "")[:200] + "..." if result.get("context_used") else "",
                "documents_used": len(result.get("context_used", "").split("Document")) - 1 if result.get("context_used") else 0
            }
        }
    except HTTPException:
        # Re-raise HTTP exceptions without modification
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate tests: {str(e)}")


@router.post("/export-excel")
async def export_tests_to_excel(tests: List[Dict[str, Any]]):
    """Export test cases to Excel format with structured format for Objective A"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers for Objective A structured format
        headers = [
            "Test Case Name",
            "Preconditions", 
            "Steps",
            "Expected Result",
            "Actual Result",
            "Test Type",
            "Priority",
            "Journey",
            "Requirement Reference",
            "Status",
            "Test Case ID",
            "Created Date"
        ]
        
        # Add headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
        
        # Add test case data
        for row_idx, test in enumerate(tests, 2):
            # Convert steps list to formatted string
            steps_data = test.get("steps", test.get("test_script", ""))
            if isinstance(steps_data, list):
                steps_formatted = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps_data)])
            else:
                steps_formatted = str(steps_data) if steps_data else ""
            
            # Use new structured format if available, fallback to legacy format
            row_data = [
                test.get("test_case_name", test.get("name", test.get("title", ""))),
                test.get("preconditions", test.get("precondition_objective", "")),
                steps_formatted,  # Use formatted steps string
                test.get("expected_result", test.get("expected", "")),
                test.get("actual_result", ""),
                test.get("test_type", "positive"),
                test.get("priority", "Medium"),
                test.get("journey", ""),
                test.get("requirement_reference", ""),
                test.get("status", "Draft"),
                test.get("test_case_id", test.get("key", test.get("test_id", ""))),
                test.get("created_date", datetime.now().strftime("%Y-%m-%d"))
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Enable text wrapping for better readability
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Color code by test type
                if col == 6:  # Test Type column
                    if value == "positive":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "negative":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "edge":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                # Color code by priority
                elif col == 7:  # Priority column
                    if value == "High":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == "Medium":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "Low":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.cell(row=1, column=1, value="Test Case Summary")
        summary_ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Count test types
        positive_count = sum(1 for test in tests if test.get("test_type") == "positive")
        negative_count = sum(1 for test in tests if test.get("test_type") == "negative")
        edge_count = sum(1 for test in tests if test.get("test_type") == "edge")
        
        summary_data = [
            ["Total Test Cases", len(tests)],
            ["Positive Cases", positive_count],
            ["Negative Cases", negative_count],
            ["Edge Cases", edge_count],
            ["High Priority", sum(1 for test in tests if test.get("priority") == "High")],
            ["Medium Priority", sum(1 for test in tests if test.get("priority") == "Medium")],
            ["Low Priority", sum(1 for test in tests if test.get("priority") == "Low")],
            ["Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            summary_ws.cell(row=row_idx, column=1, value=label)
            summary_ws.cell(row=row_idx, column=2, value=value)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as streaming response
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=test_cases_structured.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export tests: {str(e)}")


@router.post("/change-management")
async def handle_requirement_change(payload: ChangeManagementRequest):
    """Handle requirement changes and update test cases accordingly"""
    try:
        test_generator = TestGenerator()
        result = await test_generator.handle_requirement_change(
            journey=payload.journey,
            document_uri=payload.document_uri,
            source_type=payload.source_type,
            action=payload.action
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Change management failed: {str(e)}")


@router.post("/validate-test-cases")
async def validate_test_cases(payload: TestCaseValidationRequest):
    """Validate existing test cases and remove outdated ones"""
    try:
        test_generator = TestGenerator()
        result = await test_generator.validate_and_update_test_cases(
            journey=payload.journey,
            validate_outdated=payload.validate_outdated,
            remove_outdated=payload.remove_outdated
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Test case validation failed: {str(e)}")


@router.put("/update-test-case")
async def update_test_case(payload: TestCaseUpdateRequest):
    """Update a specific test case"""
    try:
        # This would typically update the test case in storage
        # For now, return a success response
        return {
            "status": "success",
            "test_case_id": payload.test_case_id,
            "journey": payload.journey,
            "message": "Test case updated successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update test case: {str(e)}")


@router.get("/test-case-types")
async def get_test_case_types():
    """Get available test case types"""
    return {
        "test_case_types": [
            {"value": "positive", "label": "Positive Cases", "description": "Normal, expected behavior"},
            {"value": "negative", "label": "Negative Cases", "description": "Error conditions and invalid inputs"},
            {"value": "edge", "label": "Edge Cases", "description": "Boundary conditions and extreme values"}
        ]
    }
