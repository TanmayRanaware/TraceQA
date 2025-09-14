"""
Tests Router with Agent-based workflow
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from ..agents.agent_manager import agent_manager
from ..schemas.tests import (
    TestGenerationRequest, 
    TestGenerationResponse,
    ChangeManagementRequest,
    TestCaseUpdateRequest,
    TestCaseValidationRequest
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io
from typing import List, Dict, Any
from datetime import datetime

router = APIRouter(prefix="/api/tests", tags=["test-generation"])

@router.post("/generate")
async def generate_tests(payload: TestGenerationRequest):
    """Generate test cases using TestGenerator directly."""
    try:
        # Debug: Check environment variables
        import os
        api_key = os.environ.get("ANTHROPIC_API_KEY", "NOT_SET")
        debug_info = f"API Key: {api_key[:20]}..." if api_key != "NOT_SET" else "API Key: NOT_SET"
        
        # Debug: Test Claude provider directly
        try:
            from ..providers.provider_factory import get_provider
            provider = get_provider()
            test_response = provider.complete([{"role": "user", "content": "Hello"}], model="claude-3-5-haiku-20241022", temperature=0.2)
            debug_info += f" | Claude test successful: {test_response[:50]}..."
        except Exception as e:
            debug_info += f" | Claude test failed: {str(e)}"
        
        # Import TestGenerator directly to bypass agent system
        from ..services.testgen import TestGenerator
        
        test_generator = TestGenerator()
        
        # Generate test cases directly
        result = await test_generator.generate_test_cases(
            journey=payload.journey,
            max_cases=payload.max_cases,
            context=payload.context or "",
            source_types=payload.source_types,
            model=payload.model,
            temperature=payload.temperature,
            context_top_k=payload.context_top_k,
            page=payload.page
        )
        
        if result.get("status") == "error":
            raise HTTPException(status_code=500, detail=result.get("message", "Test generation failed"))
        
        return {
            "journey": payload.journey,
            "tests": result.get("test_cases", []),
            "pagination": {
                "page": result.get("page", payload.page),
                "has_next_page": result.get("has_next_page", False),
                "total_pages": result.get("total_pages", 1),
                "total_available": result.get("total_available", 0)
            },
            "metadata": {
                "model_used": result.get("model_used", payload.model),
                "total_generated": result.get("total_generated", 0)
            },
            "debug": {
                "status": "completed",
                "message": "Test cases generated successfully",
                "total_generated": len(result.get("test_cases", [])),
                "context_used": result.get("context_used", "")[:200] + "..." if result.get("context_used") else "",
                "documents_used": result.get("total_available", 0),
                "claude_debug": debug_info
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Test generation failed: {str(e)}")

@router.post("/send-to-ba")
async def send_tests_to_ba(payload: TestGenerationRequest):
    """Generate and send test cases to BA using Communication Agent."""
    try:
        if not agent_manager.is_initialized():
            raise HTTPException(status_code=503, detail="Agent system not initialized")
        
        # Execute test generation workflow with email
        result = await agent_manager.execute_workflow(
            "test_generation",
            {
                "journey": payload.journey,
                "max_cases": payload.max_cases,
                "context": payload.context or "",
                "source_types": payload.source_types,
                "model": payload.model,
                "temperature": payload.temperature,
                "context_top_k": payload.context_top_k,
                "page": payload.page,
                "send_to_ba": True,
                "ba_email": payload.ba_email
            }
        )
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "Test generation and email failed"))
        
        return {
            "message": "Test cases generated and sent to BA using AI Agents",
            "workflow_id": result.get("workflow_id"),
            "journey": payload.journey,
            "ba_email": payload.ba_email,
            "test_count": len(result.get("test_result", {}).get("test_cases", []))
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send tests to BA: {str(e)}")

@router.post("/validate")
async def validate_test_cases(payload: TestCaseValidationRequest):
    """Validate test cases using Test Agent."""
    try:
        if not agent_manager.is_initialized():
            raise HTTPException(status_code=503, detail="Agent system not initialized")
        
        # Send validation task to Test Agent
        from ..agents.agent_registry import agent_registry
        result = await agent_registry.send_task_to_agent(
            "test_agent",
            "validate_test_cases",
            {
                "test_cases": payload.test_cases,
                "journey": payload.journey
            }
        )
        
        return {
            "message": "Test cases validated using AI Agent",
            "validation_result": result,
            "journey": payload.journey
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Test validation failed: {str(e)}")

@router.post("/scenarios")
async def generate_test_scenarios(payload: TestGenerationRequest):
    """Generate test scenarios using Test Agent."""
    try:
        if not agent_manager.is_initialized():
            raise HTTPException(status_code=503, detail="Agent system not initialized")
        
        # Send scenario generation task to Test Agent
        from ..agents.agent_registry import agent_registry
        result = await agent_registry.send_task_to_agent(
            "test_agent",
            "generate_test_scenarios",
            {
                "journey": payload.journey,
                "scenario_type": payload.scenario_type or "end_to_end",
                "max_scenarios": payload.max_cases or 5
            }
        )
        
        return {
            "message": "Test scenarios generated using AI Agent",
            "scenarios": result.get("scenarios", []),
            "journey": payload.journey,
            "scenario_type": payload.scenario_type or "end_to_end"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scenario generation failed: {str(e)}")

@router.post("/analyze-requirements")
async def analyze_requirements(payload: TestGenerationRequest):
    """Analyze requirements for testability using Test Agent."""
    try:
        if not agent_manager.is_initialized():
            raise HTTPException(status_code=503, detail="Agent system not initialized")
        
        # Send analysis task to Test Agent
        from ..agents.agent_registry import agent_registry
        result = await agent_registry.send_task_to_agent(
            "test_agent",
            "analyze_requirements",
            {
                "journey": payload.journey,
                "requirements_text": payload.context or ""
            }
        )
        
        return {
            "message": "Requirements analyzed using AI Agent",
            "analysis": result.get("analysis", {}),
            "journey": payload.journey
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Requirements analysis failed: {str(e)}")

@router.get("/export/{journey}")
async def export_tests_to_excel(journey: str, max_cases: int = 50):
    """Export test cases to Excel using Test Agent."""
    try:
        if not agent_manager.is_initialized():
            raise HTTPException(status_code=503, detail="Agent system not initialized")
        
        # Generate test cases first
        result = await agent_manager.execute_workflow(
            "test_generation",
            {
                "journey": journey,
                "max_cases": max_cases,
                "send_to_ba": False
            }
        )
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "Test generation failed"))
        
        test_cases = result.get("test_result", {}).get("test_cases", [])
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{journey} Test Cases"
        
        # Headers
        headers = ["Test Case Name", "Preconditions", "Test Steps", "Expected Result", "Priority", "Category", "Test Data"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add test cases
        for row, test_case in enumerate(test_cases, 2):
            ws.cell(row=row, column=1, value=test_case.get("test_case_name", ""))
            ws.cell(row=row, column=2, value=test_case.get("preconditions", ""))
            ws.cell(row=row, column=3, value="\n".join(test_case.get("test_steps", [])))
            ws.cell(row=row, column=4, value=test_case.get("expected_result", ""))
            ws.cell(row=row, column=5, value=test_case.get("priority", ""))
            ws.cell(row=row, column=6, value=test_case.get("category", ""))
            ws.cell(row=row, column=7, value=str(test_case.get("test_data", {})))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={journey}_test_cases.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.get("/agent-status")
async def get_test_agent_status():
    """Get status of Test Agent."""
    try:
        if not agent_manager.is_initialized():
            return {"success": False, "error": "Agent system not initialized"}
        
        from ..agents.agent_registry import agent_registry
        test_agent = agent_registry.get_agent("test_agent")
        
        if not test_agent:
            return {"success": False, "error": "Test Agent not found"}
        
        return {
            "success": True,
            "agent_status": test_agent.get_status(),
            "message": "Test Agent status retrieved"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get test agent status: {str(e)}")
